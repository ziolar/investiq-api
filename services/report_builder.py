import io
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, GradientFill
)
from openpyxl.utils import get_column_letter


def _border(style="thin"):
    s = Side(style=style, color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)


def build_excel_report(analysis: dict) -> bytes:
    wb = Workbook()

    _sheet_summary(wb, analysis)
    _sheet_scoring(wb, analysis)
    _sheet_industry(wb, analysis)
    _sheet_brief(wb, analysis)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _sheet_summary(wb, analysis):
    ws = wb.active
    ws.title = "投资摘要"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20

    scoring = analysis.get("scoring", {})
    total = scoring.get("total_pct", 0)
    rec = scoring.get("recommendation", "N/A")
    rec_color = scoring.get("recommendation_color", "#6b7280").lstrip("#")

    # Title
    ws.merge_cells("A1:D1")
    c = ws["A1"]
    c.value = "一级市场投资评估报告"
    c.font = Font(name="微软雅黑", size=18, bold=True, color="1E3A5F")
    c.fill = _fill("EBF3FB")
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    # Basic info
    info = [
        ("项目名称", analysis.get("project_name", "N/A")),
        ("公司名称", analysis.get("company_name", "N/A")),
        ("所处行业", analysis.get("industry", "N/A")),
        ("融资阶段", analysis.get("stage", "N/A")),
    ]
    for i, (label, value) in enumerate(info, start=2):
        ws.row_dimensions[i].height = 22
        a = ws.cell(row=i, column=1, value=label)
        a.font = Font(name="微软雅黑", bold=True, size=10, color="4B5563")
        a.fill = _fill("F9FAFB")
        a.border = _border()
        a.alignment = Alignment(vertical="center")
        b = ws.cell(row=i, column=2, value=value)
        b.font = Font(name="微软雅黑", size=10)
        b.border = _border()
        b.alignment = Alignment(vertical="center", wrap_text=True)
        ws.merge_cells(f"B{i}:D{i}")

    # Score & recommendation
    ws.row_dimensions[7].height = 30
    ws.merge_cells("A7:D7")
    ws["A7"].value = f"综合评分：{total:.1f} / 100   投资建议：{rec}"
    ws["A7"].font = Font(name="微软雅黑", size=14, bold=True, color="FFFFFF")
    ws["A7"].fill = _fill(rec_color)
    ws["A7"].alignment = Alignment(horizontal="center", vertical="center")

    # Dimension score bar
    ws.row_dimensions[9].height = 20
    ws.cell(row=9, column=1, value="评估维度").font = Font(bold=True, size=10, name="微软雅黑")
    ws.cell(row=9, column=2, value="原始评分（/10）").font = Font(bold=True, size=10, name="微软雅黑")
    ws.cell(row=9, column=3, value="权重").font = Font(bold=True, size=10, name="微软雅黑")
    ws.cell(row=9, column=4, value="加权得分").font = Font(bold=True, size=10, name="微软雅黑")
    for col in range(1, 5):
        ws.cell(row=9, column=col).fill = _fill("1E3A5F")
        ws.cell(row=9, column=col).font = Font(bold=True, size=10, name="微软雅黑", color="FFFFFF")
        ws.cell(row=9, column=col).alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=9, column=col).border = _border()

    dims = scoring.get("dimensions", {})
    for idx, (dim, data) in enumerate(dims.items(), start=10):
        ws.row_dimensions[idx].height = 20
        ws.cell(row=idx, column=1, value=dim).border = _border()
        ws.cell(row=idx, column=1).font = Font(name="微软雅黑", size=10)
        ws.cell(row=idx, column=1).alignment = Alignment(vertical="center")

        raw = ws.cell(row=idx, column=2, value=data["raw_score"])
        raw.border = _border()
        raw.alignment = Alignment(horizontal="center", vertical="center")
        raw.font = Font(name="微软雅黑", size=10)

        w = ws.cell(row=idx, column=3, value=f"{int(data['weight']*100)}%")
        w.border = _border()
        w.alignment = Alignment(horizontal="center", vertical="center")
        w.font = Font(name="微软雅黑", size=10)

        weighted_val = round(data["raw_score"] * data["weight"] * 10, 1)
        wt = ws.cell(row=idx, column=4, value=weighted_val)
        wt.border = _border()
        wt.alignment = Alignment(horizontal="center", vertical="center")
        wt.font = Font(name="微软雅黑", size=10, bold=True)

        if idx % 2 == 0:
            for col in range(1, 5):
                ws.cell(row=idx, column=col).fill = _fill("F3F4F6")

    # Highlights & risks
    row = 10 + len(dims) + 1
    ws.cell(row=row, column=1, value="核心亮点").font = Font(bold=True, size=10, name="微软雅黑", color="16A34A")
    for h in analysis.get("key_highlights", []):
        row += 1
        ws.cell(row=row, column=1, value=f"• {h}").font = Font(size=9, name="微软雅黑")
        ws.merge_cells(f"A{row}:D{row}")
        ws.row_dimensions[row].height = 18

    row += 1
    ws.cell(row=row, column=1, value="主要风险").font = Font(bold=True, size=10, name="微软雅黑", color="DC2626")
    for r in analysis.get("key_risks", []):
        row += 1
        ws.cell(row=row, column=1, value=f"• {r}").font = Font(size=9, name="微软雅黑")
        ws.merge_cells(f"A{row}:D{row}")
        ws.row_dimensions[row].height = 18


def _sheet_scoring(wb, analysis):
    ws = wb.create_sheet("详细评分")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 22
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 50

    headers = ["评估维度", "评估子项", "得分(/10)", "子项权重", "依据说明"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = Font(bold=True, size=10, name="微软雅黑", color="FFFFFF")
        c.fill = _fill("1E3A5F")
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = _border()
    ws.row_dimensions[1].height = 22

    row = 2
    framework = analysis.get("framework", {})
    dims = analysis.get("scoring", {}).get("dimensions", {})

    for dim, dim_data in dims.items():
        sub_count = len(dim_data.get("subcriteria", {}))
        start_row = row
        fw_sub = framework.get(dim, {}).get("subcriteria", {})

        for sub, sub_data in dim_data.get("subcriteria", {}).items():
            ws.row_dimensions[row].height = 30
            c = ws.cell(row=row, column=2, value=sub)
            c.font = Font(size=9, name="微软雅黑")
            c.border = _border()
            c.alignment = Alignment(vertical="center", wrap_text=True)

            score_val = sub_data.get("score", 0)
            sc = ws.cell(row=row, column=3, value=score_val)
            sc.font = Font(size=10, bold=True, name="微软雅黑")
            sc.border = _border()
            sc.alignment = Alignment(horizontal="center", vertical="center")
            # Color by score
            if score_val >= 8:
                sc.font = Font(size=10, bold=True, name="微软雅黑", color="16A34A")
            elif score_val <= 4:
                sc.font = Font(size=10, bold=True, name="微软雅黑", color="DC2626")

            sw = fw_sub.get(sub, 0)
            wc = ws.cell(row=row, column=4, value=f"{int(sw*100)}%")
            wc.font = Font(size=9, name="微软雅黑")
            wc.border = _border()
            wc.alignment = Alignment(horizontal="center", vertical="center")

            ev = ws.cell(row=row, column=5, value=sub_data.get("evidence", ""))
            ev.font = Font(size=9, name="微软雅黑")
            ev.border = _border()
            ev.alignment = Alignment(vertical="center", wrap_text=True)

            row += 1

        # Merge dimension column
        ws.merge_cells(f"A{start_row}:A{row-1}")
        dc = ws[f"A{start_row}"]
        dc.value = dim
        dc.font = Font(bold=True, size=10, name="微软雅黑")
        dc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        dc.border = _border()
        dc.fill = _fill("EBF3FB")


def _sheet_industry(wb, analysis):
    ws = wb.create_sheet("行业数据")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 25
    ws.column_dimensions["B"].width = 55

    industry_data = analysis.get("industry_data", {})
    if not industry_data:
        ws["A1"] = "暂无行业数据"
        return

    ws.merge_cells("A1:B1")
    ws["A1"].value = f"行业基准数据：{industry_data.get('label', '')}"
    ws["A1"].font = Font(bold=True, size=13, name="微软雅黑", color="1E3A5F")
    ws["A1"].fill = _fill("EBF3FB")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    fields = [
        ("典型增长率", "typical_arr_growth"),
        ("典型估值倍数", "typical_valuation_multiple"),
        ("A轮融资规模", "median_series_a_size"),
        ("毛利率基准", "benchmark_gross_margin"),
        ("典型退出倍数", "typical_exit_multiple"),
    ]
    for i, (label, key) in enumerate(fields, start=2):
        ws.row_dimensions[i].height = 20
        a = ws.cell(row=i, column=1, value=label)
        a.font = Font(bold=True, size=10, name="微软雅黑")
        a.fill = _fill("F9FAFB")
        a.border = _border()
        b = ws.cell(row=i, column=2, value=industry_data.get(key, "N/A"))
        b.font = Font(size=10, name="微软雅黑")
        b.border = _border()
        b.alignment = Alignment(wrap_text=True, vertical="center")

    row = len(fields) + 3
    for list_field, label in [
        ("key_metrics", "关键指标"), ("key_success_factors", "成功要素"),
        ("key_risks", "主要风险"), ("exit_cases", "退出案例")
    ]:
        items = industry_data.get(list_field, [])
        if not items:
            continue
        ws.row_dimensions[row].height = 20
        ws.merge_cells(f"A{row}:B{row}")
        ws[f"A{row}"].value = label
        ws[f"A{row}"].font = Font(bold=True, size=10, name="微软雅黑", color="1E3A5F")
        ws[f"A{row}"].fill = _fill("DBEAFE")
        ws[f"A{row}"].border = _border()
        row += 1
        for item in items:
            ws.row_dimensions[row].height = 18
            ws.merge_cells(f"A{row}:B{row}")
            ws[f"A{row}"].value = f"• {item}"
            ws[f"A{row}"].font = Font(size=9, name="微软雅黑")
            ws[f"A{row}"].border = _border()
            row += 1


def _sheet_brief(wb, analysis):
    ws = wb.create_sheet("投资简报")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 80

    ws.merge_cells("A1:A1")
    ws["A1"].value = "投资简报"
    ws["A1"].font = Font(bold=True, size=14, name="微软雅黑", color="1E3A5F")
    ws["A1"].fill = _fill("EBF3FB")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    brief = analysis.get("investment_brief", "暂无投资简报")
    ws["A2"].value = brief
    ws["A2"].font = Font(size=11, name="微软雅黑")
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="top")
    ws.row_dimensions[2].height = 300
